from openpyxl import load_workbook
from openpyxl import Workbook
import datetime
import argparse
import os

# Fonction principale
def main(input_file):
    # Charger le fichier Excel existant
    workbook = load_workbook(filename=input_file)
    sheet = workbook.active

    # Dictionnaire pour stocker les données
    Dict = {}

    cpt = 0
    # Lire les données de la feuille
    for row in sheet.iter_rows(min_row=2, values_only=True):
        cpt += 1
        matricule = row[0]
        # Sélectionner les colonnes d'intérêt
        row_data = [str(row[i]) for i in [1, 29, 30, 25, 23, 28, 21, 56]]
        if isinstance(row[30], datetime.datetime):
            row_data[2] = str(row[30].date())

        if matricule not in Dict:
            Dict[matricule] = row_data
        else:
            Dict[matricule] += row_data  

    print(f"Nombre de lignes traités sans la 1ere qui contient le nom des colonnes : {cpt}")
    print(f"Nombre de matricules traités : {len(Dict.keys())}")

    # Créer un nouveau classeur et une nouvelle feuille
    new_workbook = Workbook()
    new_sheet = new_workbook.active

    # Écrire les données dans la feuille
    for i, key in enumerate(Dict.keys(), start=1):  # Parcourir toutes les clés
        for j, value in enumerate(Dict[key], start=1):  # Parcourir les valeurs associées à chaque clé
            new_sheet.cell(row=i, column=j, value=value)  # Écrire chaque valeur dans la colonne correspondante

    # Déterminer le chemin vers le dossier Téléchargements
    downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")

    # Extraire le nom sans l'extension
    base_name = os.path.splitext(os.path.basename(input_file))[0]
    output_file = os.path.join(downloads_path, f'{base_name}_transformed.xlsx')

    # Enregistrer le nouveau fichier Excel
    new_workbook.save(output_file)
    print(f"Les données ont été écrites dans {output_file} avec succès.")

if __name__ == "__main__":
    # Configurer argparse pour gérer les arguments de ligne de commande
    parser = argparse.ArgumentParser(description='Transforme un fichier Excel en un format spécifique.')
    parser.add_argument('input_file', type=str, help='Le chemin du fichier Excel à transformer.')

    args = parser.parse_args()
    
    # Appeler la fonction principale avec le fichier d'entrée
    main(args.input_file)
